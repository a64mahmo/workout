"""
Seed all programs from the Excel spreadsheets in the programs/ folder.

Parses each .xlsx file, extracts weeks / sessions / exercises and creates:
  - Plan (one per file)
  - PlanSession (one per week × session)
  - PlanExercise (one per exercise row)
  - Exercise records (created globally if they don't exist yet)

Usage:
  # Seed for a specific user by email
  python seed_xlsx_programs.py --email user@example.com

  # Seed for a specific user by ID
  python seed_xlsx_programs.py --user-id <uuid>

  # List what plans would be created without writing to the DB
  python seed_xlsx_programs.py --dry-run
"""
import asyncio
import argparse
import os
import re
import sys
import uuid
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
load_dotenv(Path(__file__).parent / ".env")

import openpyxl
from sqlalchemy import select

from app.database import async_session, init_db
from app.models.models import Exercise, Plan, PlanExercise, PlanSession, User

PROGRAMS_DIR = Path(__file__).parent.parent / "programs"

# Maps filename stem -> human-readable plan name
PLAN_NAMES = {
    "POWERBUILDING 2.0 SPREADSHEET 4x":    "Powerbuilding 2.0 (4x/Week)",
    "POWERBUILDING 2.0 SPREADSHEET 5-6x":  "Powerbuilding 2.0 (5-6x/Week)",
    "PowerBuilding-3.0-5x-vvujkv":         "Powerbuilding 3.0 (5x/Week)",
    "POWERBUILDING-4x-Spreadsheet":        "Powerbuilding System (4x/Week)",
    "POWERBUILDING-6x-Spreadsheet":        "Powerbuilding System (5-6x/Week)",
}

# ─── Parsing helpers ──────────────────────────────────────────────────────────

def _clean_name(raw: str) -> str:
    """Strip superset labels, bracketed alternatives, and common suffixes."""
    s = str(raw).strip()
    s = re.sub(r'^[A-Z]\d+:\s*', '', s)               # "A1: ", "B2: " …
    s = re.sub(r'\s*\[.*?\]', '', s)                    # "[or Nordic ham curl]"
    s = re.sub(r'\s*\(or[^)]*\)', '', s, flags=re.I)    # "(or box squat)"
    s = re.sub(r'\s*\(choice\)', '', s, flags=re.I)
    s = re.sub(r'\s*\(optional\)', '', s, flags=re.I)
    return s.strip()


def _parse_reps(raw) -> int | None:
    """Return a usable integer rep target from whatever Excel gives us."""
    if raw is None:
        return None
    # Excel mis-reads "4/6" as a date (April 6th) → use the day component
    if isinstance(raw, datetime):
        return raw.day
    s = str(raw).strip().strip('"').strip("'")
    if s.upper() in ("AMRAP", "N/A", ""):
        return 0
    # "8-10each" / "10-12" / "12-15" → take the first number
    m = re.match(r'(\d+)', s)
    return int(m.group(1)) if m else 0


def _parse_sets(raw) -> int:
    if raw is None:
        return 1
    s = str(raw).strip()
    m = re.match(r'(\d+)', s)
    return int(m.group(1)) if m else 1


def _parse_weight(raw) -> float | None:
    """Convert %1RM column to a float (0–100 scale)."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        v = float(raw)
        # Values stored as decimals (0.8) need converting; whole numbers are already %
        return round(v * 100, 1) if v <= 1.0 else v
    s = str(raw).strip()
    if s.upper() == "N/A" or s == "":
        return None
    # "75-80%" → midpoint 77.5; "82.5-87.5%" → 85.0
    m = re.findall(r'[\d.]+', s)
    if len(m) >= 2:
        return round((float(m[0]) + float(m[1])) / 2, 1)
    if len(m) == 1:
        v = float(m[0])
        return round(v * 100, 1) if v <= 1.0 else v
    return None


def _parse_rpe(raw) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        # "7/8" mis-read as July 8 → use month as lower bound
        return float(raw.month)
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip().strip('"').strip("'")
    if s.upper() in ("N/A", ""):
        return None
    # "~6-8" / "6-8" → midpoint
    s = s.lstrip("~")
    m = re.findall(r'[\d.]+', s)
    if len(m) >= 2:
        return round((float(m[0]) + float(m[1])) / 2, 1)
    if len(m) == 1:
        return float(m[0])
    return None


_REST_MAP = {
    "0": 0, "0min": 0, "0 min": 0,
    "1-2 min": 90,  "1-2min": 90,
    "2-3 min": 150, "2-3min": 150,
    "2-4 min": 180, "2-4min": 180,
    "3-4 min": 180, "3-4min": 180,
    "3-5 min": 240, "3-5min": 240,
    "4-5 min": 240, "4-5min": 240,
    "4-6 min": 300, "4-6min": 300,
    "5-6 min": 300, "5-6min": 300,
    "5-7 min": 360, "5-7min": 360,
}

def _parse_rest(raw) -> int:
    if raw is None:
        return 90
    s = str(raw).strip().lower()
    if s in _REST_MAP:
        return _REST_MAP[s]
    # Fallback: extract first number and multiply by 60
    m = re.match(r'(\d+)', s)
    return int(m.group(1)) * 60 if m else 90


# ─── Spreadsheet parser ───────────────────────────────────────────────────────

def _week_num(week_label: str) -> int:
    """Extract integer week number from labels like 'Week 1', 'Week 8 - semi deload', 'Week 10A'."""
    m = re.search(r'\d+', str(week_label))
    return int(m.group()) if m else 1


def _is_skip_row(cell_b) -> bool:
    """True for REST DAY / informational rows we should ignore."""
    if cell_b is None:
        return False
    s = str(cell_b).upper()
    return (
        s.startswith("REST DAY")
        or s.startswith("IF YOU HAVE")
        or s.startswith("OPTIONAL")
        or s.startswith("NOTE")
        or s.startswith("IMPORTANT")
        or s.startswith("*")
    )


def parse_xlsx(path: Path) -> dict:
    """
    Returns:
        {
          "plan_name": str,
          "description": str,
          "sessions": [
              {
                "name": "Week 1 - FULL BODY 1",
                "order": int,
                "exercises": [
                    {
                      "name": str,
                      "warmup_sets": int,
                      "target_sets": int,
                      "target_reps": int,
                      "target_weight": float | None,
                      "target_rpe": float | None,
                      "rest_seconds": int,
                      "notes": str | None,
                    }, ...
                ]
              }, ...
          ]
        }
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    stem = path.stem
    plan_name = PLAN_NAMES.get(stem, stem.replace("-", " ").title())

    # Detect units from row 4 col D
    units = "lbs"
    for row in ws.iter_rows(min_row=1, max_row=10, max_col=5, values_only=True):
        if row[3] in ("kg", "lbs"):
            units = row[3]
            break

    description = f"Units: {units}. 1RM: Squat=100, Bench=100, Deadlift=100, OHP=100"

    sessions = []
    current_week = ""
    current_session_name = ""
    current_exercises: list[dict] = []
    order_idx = 0
    in_program = False  # only True after the first "Week N / Exercise" header row

    def flush_session():
        nonlocal current_exercises
        if current_session_name and current_exercises:
            sessions.append({
                "name": f"{current_week} - {current_session_name}",
                "week_number": _week_num(current_week),
                "order": order_idx,
                "exercises": current_exercises,
            })
        current_exercises = []

    for row in ws.iter_rows(min_row=1, values_only=True):
        # Columns: A=0, B=1, C=2, D=3, E=4, F=5, G=6, H=7, I=8, J=9, K=10
        col_b = row[1] if len(row) > 1 else None
        col_c = row[2] if len(row) > 2 else None
        col_d = row[3] if len(row) > 3 else None  # warmup sets
        col_e = row[4] if len(row) > 4 else None  # working sets
        col_f = row[5] if len(row) > 5 else None  # reps
        col_h = row[7] if len(row) > 7 else None  # %1RM  (used as target_weight)
        col_i = row[8] if len(row) > 8 else None  # RPE
        col_j = row[9] if len(row) > 9 else None  # rest
        col_k = row[10] if len(row) > 10 else None  # notes

        # ── Week header row (e.g. col_b="Week 1", col_c="Exercise") ──
        if col_b and str(col_b).startswith("Week") and col_c == "Exercise":
            flush_session()
            current_week = str(col_b).strip()
            current_session_name = ""
            in_program = True
            continue

        if not in_program:
            continue

        if _is_skip_row(col_b):
            continue

        # ── Session start (col_b has session name, col_c has first exercise) ──
        if col_b and not str(col_b).startswith("Week"):
            flush_session()
            order_idx += 1
            # Normalize session name: strip trailing ":", extra descriptor after ":"
            raw_session = str(col_b).strip().rstrip(":")
            # "Full Body 1: Squat, OHP" → keep just "Full Body 1"
            raw_session = raw_session.split(":")[0].strip()
            current_session_name = raw_session
            current_exercises = []

            # The first exercise is on the same row as the session name
            if col_c and col_c != "Exercise":
                ex_name = _clean_name(str(col_c))
                if ex_name:
                    current_exercises.append({
                        "name": ex_name,
                        "warmup_sets": _parse_sets(col_d) if col_d else 0,
                        "target_sets": _parse_sets(col_e),
                        "target_reps": _parse_reps(col_f),
                        "target_weight": _parse_weight(col_h),
                        "target_rpe": _parse_rpe(col_i),
                        "rest_seconds": _parse_rest(col_j),
                        "notes": str(col_k).strip() if col_k else None,
                    })
            continue

        # ── Exercise continuation row (col_b is None, col_c has exercise) ──
        if col_b is None and col_c and col_c != "Exercise" and current_session_name:
            ex_name = _clean_name(str(col_c))
            if ex_name:
                current_exercises.append({
                    "name": ex_name,
                    "warmup_sets": _parse_sets(col_d) if col_d else 0,
                    "target_sets": _parse_sets(col_e),
                    "target_reps": _parse_reps(col_f),
                    "target_weight": _parse_weight(col_h),
                    "target_rpe": _parse_rpe(col_i),
                    "rest_seconds": _parse_rest(col_j),
                    "notes": str(col_k).strip() if col_k else None,
                })

    flush_session()

    return {"plan_name": plan_name, "description": description, "sessions": sessions}


# ─── DB seeding ───────────────────────────────────────────────────────────────

async def seed_plan_for_user(user_id: str, plan_data: dict, db) -> str:
    """
    Insert one parsed plan for the given user.
    Returns a status string for logging.
    """
    plan_name = plan_data["plan_name"]

    # Skip if plan already exists for this user
    existing = await db.execute(
        select(Plan).where(Plan.name == plan_name, Plan.user_id == user_id)
    )
    if existing.scalar_one_or_none():
        return f"  Skipped '{plan_name}' — already exists"

    # Ensure all exercises exist; build name→id map
    exercise_map: dict[str, str] = {}
    all_ex_names = {
        ex["name"]
        for s in plan_data["sessions"]
        for ex in s["exercises"]
        if ex["name"]
    }
    for ex_name in all_ex_names:
        result = await db.execute(select(Exercise).where(Exercise.name == ex_name))
        existing_ex = result.scalar_one_or_none()
        if existing_ex:
            exercise_map[ex_name] = existing_ex.id
        else:
            # Guess muscle group from name — users can update later
            muscle = _guess_muscle(ex_name)
            new_ex = Exercise(
                id=str(uuid.uuid4()),
                name=ex_name,
                muscle_group=muscle,
                description="",
            )
            db.add(new_ex)
            exercise_map[ex_name] = new_ex.id

    await db.flush()

    # Create the plan
    plan = Plan(
        id=str(uuid.uuid4()),
        user_id=user_id,
        name=plan_name,
        description=plan_data["description"],
    )
    db.add(plan)
    await db.flush()

    # Create sessions + exercises
    for session_data in plan_data["sessions"]:
        ps = PlanSession(
            id=str(uuid.uuid4()),
            plan_id=plan.id,
            name=session_data["name"],
            week_number=session_data["week_number"],
            order_index=session_data["order"],
        )
        db.add(ps)
        await db.flush()

        for ex_idx, ex in enumerate(session_data["exercises"]):
            if ex["name"] not in exercise_map:
                continue
            pe = PlanExercise(
                id=str(uuid.uuid4()),
                plan_session_id=ps.id,
                exercise_id=exercise_map[ex["name"]],
                order_index=ex_idx,
                target_sets=ex["target_sets"],
                target_reps=ex["target_reps"] or 0,
                target_weight=ex["target_weight"],
                target_rpe=ex["target_rpe"],
                rest_seconds=ex["rest_seconds"],
                notes=ex["notes"],
            )
            db.add(pe)

    await db.commit()
    session_count = len(plan_data["sessions"])
    return f"  Created '{plan_name}' ({session_count} sessions)"


_MUSCLE_KEYWORDS = {
    "squat": "legs", "deadlift": "back", "bench": "chest", "press": "chest",
    "overhead": "shoulders", "ohp": "shoulders", "row": "back", "pull": "back",
    "curl": "biceps", "tricep": "triceps", "extension": "triceps",
    "calf": "legs", "leg": "legs", "hip": "glutes", "glute": "glutes",
    "hamstring": "hamstrings", "rdl": "hamstrings", "ham": "hamstrings",
    "lateral": "shoulders", "face pull": "shoulders", "shrug": "traps",
    "plank": "core", "crunch": "core", "ab": "core", "sit": "core",
    "neck": "neck", "back extension": "back", "flye": "chest", "fly": "chest",
    "dip": "chest", "push": "chest",
}

def _guess_muscle(name: str) -> str:
    lower = name.lower()
    for keyword, muscle in _MUSCLE_KEYWORDS.items():
        if keyword in lower:
            return muscle
    return "other"


# ─── CLI ──────────────────────────────────────────────────────────────────────

async def run(args: argparse.Namespace) -> None:
    xlsx_files = sorted(PROGRAMS_DIR.glob("*.xlsx"))
    if not xlsx_files:
        print(f"No .xlsx files found in {PROGRAMS_DIR}")
        sys.exit(1)

    # Parse all files first
    parsed = []
    for f in xlsx_files:
        print(f"Parsing {f.name}...")
        data = parse_xlsx(f)
        parsed.append(data)
        print(f"  → '{data['plan_name']}': {len(data['sessions'])} sessions")

    if args.dry_run:
        print("\nDry run — nothing written to the database.")
        return

    await init_db()

    # Resolve user
    async with async_session() as db:
        if args.email:
            result = await db.execute(select(User).where(User.email == args.email))
        else:
            result = await db.execute(select(User).where(User.id == args.user_id))
        user = result.scalar_one_or_none()

    if not user:
        identifier = args.email or args.user_id
        print(f"Error: No user found for '{identifier}'")
        sys.exit(1)

    print(f"\nSeeding programs for {user.email} ({user.id})...")
    for plan_data in parsed:
        async with async_session() as db:
            status = await seed_plan_for_user(user.id, plan_data, db)
        print(status)

    print("\nDone.")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Seed workout programs from programs/*.xlsx for a user"
    )
    target = parser.add_mutually_exclusive_group(required=True)
    target.add_argument("--email", help="Target user's email address")
    target.add_argument("--user-id", help="Target user's UUID")
    target.add_argument("--dry-run", action="store_true",
                        help="Parse files and print plan names without writing to DB")
    args = parser.parse_args()

    asyncio.run(run(args))


if __name__ == "__main__":
    main()
